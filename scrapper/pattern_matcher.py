"""
Pattern Matcher - Watches race data and results, creates Excel with highlighted winners

This script:
1. Watches races.csv and race_results.csv for changes
2. When new results are found, matches winners to race data by horse name
3. Creates/updates an Excel file with race data and highlighted winners

The matching is done by finding the horse name from OCR results in the race data.
OCR results have combined name+jockey, so we extract just the horse name for matching.

Requirements:
    pip install openpyxl watchdog
"""

import os
import csv
import time
import re
from datetime import datetime
from pathlib import Path
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Configuration
RACES_CSV = Path(__file__).parent / "races.csv"
RESULTS_CSV = Path(__file__).parent / "race_results.csv"
OUTPUT_EXCEL = Path(__file__).parent / "race_analysis.xlsx"

# Highlight colors for positions
POSITION_COLORS = {
    1: PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),  # Gold for 1st
    2: PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid"),  # Silver for 2nd
    3: PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid"),  # Bronze for 3rd
    4: PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # Light green for 4th
}


def normalize_name(name):
    """Normalize horse name for matching (uppercase, remove extra spaces)."""
    if not name:
        return ""
    # Remove any non-alphanumeric except spaces, convert to uppercase
    name = re.sub(r'[^A-Za-z0-9\s]', '', name)
    return ' '.join(name.upper().split())


def extract_horse_name_from_ocr(ocr_text):
    """
    Extract horse name from OCR text which contains name + jockey combined.
    Horse names are typically UPPERCASE, jockey names are mixed case.
    """
    if not ocr_text:
        return ""
    
    # Split into words
    words = ocr_text.split()
    
    # Find consecutive uppercase words (horse name)
    horse_words = []
    for word in words:
        # Check if word is mostly uppercase (horse name part)
        upper_count = sum(1 for c in word if c.isupper())
        if len(word) > 0 and upper_count >= len(word) * 0.7:
            horse_words.append(word.upper())
        else:
            # Stop when we hit lowercase (jockey name starts)
            if horse_words:
                break
    
    return ' '.join(horse_words)


def load_races():
    """Load race data from races.csv."""
    if not RACES_CSV.exists():
        return []
    
    races = []
    with open(RACES_CSV, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            races.append(row)
    return races


def load_results():
    """Load race results from race_results.csv."""
    if not RESULTS_CSV.exists():
        return []
    
    results = []
    with open(RESULTS_CSV, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            results.append(row)
    return results


def extract_date_from_scraped_at(scraped_at):
    """
    Extract date (YYYY-MM-DD) from scraped_at field.
    Handles formats:
    - "2025-12-13 21:00:00" (races.csv format)
    - "2025-12-13T21:00:00.123456" (race_results.csv ISO format)
    """
    if not scraped_at:
        return ""
    try:
        # Handle both space and T separator
        if 'T' in scraped_at:
            return scraped_at.split('T')[0]
        else:
            return scraped_at.split(' ')[0]
    except:
        return ""


def get_results_for_race(race_row, all_results):
    """
    Get results that belong to a specific race.
    Uses video_race_time from results and race_time_uk from race for matching.
    Also checks that both are from the same date (race times repeat every 24h).
    
    Matching logic:
    1. Same date (from scraped_at)
    2. video_race_time == race_time_uk
    """
    race_time_uk = race_row.get('race_time_uk', '')
    if not race_time_uk:
        return []
    
    # Get date from race data
    race_date = extract_date_from_scraped_at(race_row.get('scraped_at', ''))
    
    # Find all results that match this race time AND date
    matching_results = []
    for result in all_results:
        # First check: same date
        result_date = extract_date_from_scraped_at(result.get('scraped_at', ''))
        if race_date and result_date and race_date != result_date:
            continue  # Different days, skip
        
        # Second check: matching time
        video_time = result.get('video_race_time', '')
        
        if video_time:
            # Direct match: video_race_time should equal race_time_uk
            if video_time == race_time_uk:
                matching_results.append(result)
        else:
            # Fallback: use timestamp-based matching
            result_time = result.get('race_time', '')
            if '_' in result_time:
                try:
                    time_part = result_time.split('_')[1]
                    result_hour = time_part[0:2]
                    result_min = time_part[2:4]
                    
                    race_parts = race_time_uk.split(':')
                    if len(race_parts) == 2:
                        race_minutes = int(race_parts[0]) * 60 + int(race_parts[1])
                        result_minutes = int(result_hour) * 60 + int(result_min)
                        
                        if 0 <= result_minutes - race_minutes <= 3:
                            matching_results.append(result)
                except:
                    pass
    
    return matching_results


def match_winners_to_race(race_row, all_results):
    """
    Match race results to runners in a race.
    Uses horse_number from results to directly map to runner in race.
    
    Matching logic:
    - horse_number in results (e.g., 1, 5, 3) matches to name_1, name_5, name_3 in races.csv
    - This is more reliable than name matching since OCR can garble names
    
    Returns a dict mapping runner number to position (1st, 2nd, 3rd, 4th).
    """
    winner_positions = {}
    
    # Get results specific to this race (by time)
    results = get_results_for_race(race_row, all_results)
    
    if not results:
        return winner_positions
    
    # Get list of valid runner numbers from this race
    valid_runner_nums = set()
    for key in race_row.keys():
        if key.startswith('name_'):
            runner_num = key.split('_')[1]
            if race_row.get(key):  # Only if name is not empty
                valid_runner_nums.add(runner_num)
    
    # Match by horse_number directly
    for result in results:
        position = int(result.get('position', 0))
        if position < 1 or position > 4:
            continue
        
        # Get horse number from result (the number shown in video results)
        horse_number = str(result.get('horse_number', '')).strip()
        
        # If horse_number exists in race, that's our match!
        if horse_number and horse_number in valid_runner_nums:
            winner_positions[horse_number] = position
    
    return winner_positions


def get_race_signature(race_row):
    """
    Generate a unique signature for a race based on horses and jockeys.
    Signature includes: horse name + jockey name (sorted alphabetically).
    Ignores runner numbers and odds - same horses with same jockeys = duplicate.
    """
    # Collect all horse:jockey pairs
    runner_pairs = []
    for i in range(1, 20):  # Max 20 runners
        name = race_row.get(f'name_{i}', '')
        jockey = race_row.get(f'jockey_{i}', '')
        if name:  # Only if runner exists
            runner_pairs.append(f"{name}:{jockey}")
    
    # Sort alphabetically so order doesn't matter, then join
    return "|".join(sorted(runner_pairs))


def create_excel_with_highlights():
    """Create Excel file with race data and highlighted winners."""
    races = load_races()
    results = load_results()
    
    if not races:
        print("No race data found in races.csv")
        return
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Race Analysis"
    
    # Get original headers from first race
    original_headers = list(races[0].keys())
    
    # Build new headers: date first, then original, then duplicate detection columns
    headers = ['date'] + original_headers + ['duplicate_count', 'last_seen']
    
    # Pre-compute race signatures and track duplicates
    race_signatures = {}  # signature -> list of (date_time, row_idx)
    for idx, race in enumerate(races):
        sig = get_race_signature(race)
        race_date = extract_date_from_scraped_at(race.get('scraped_at', ''))
        race_time = race.get('race_time', '')  # Get race_time for display
        date_time = f"{race_date} {race_time}" if race_date and race_time else race_date
        if sig not in race_signatures:
            race_signatures[sig] = []
        race_signatures[sig].append((date_time, idx))
    
    # Write headers with styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Write race data
    for row_idx, race in enumerate(races, 2):
        # Match winners for this race
        winner_positions = match_winners_to_race(race, results)
        
        # Get race date and signature for duplicate detection
        race_date = extract_date_from_scraped_at(race.get('scraped_at', ''))
        signature = get_race_signature(race)
        
        # Find duplicate info
        sig_occurrences = race_signatures.get(signature, [])
        duplicate_count = len(sig_occurrences)
        
        # Find last occurrence before this one (shows date + race_time)
        current_idx = row_idx - 2  # Convert back to 0-based index
        earlier_times = [dt for dt, i in sig_occurrences if i < current_idx]
        last_seen = earlier_times[-1] if earlier_times else ""
        
        col_idx = 1
        
        # Write date column first
        cell = ws.cell(row=row_idx, column=col_idx, value=race_date)
        col_idx += 1
        
        # Write original data
        for header in original_headers:
            value = race.get(header, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Check if this is a name column and if runner is a winner
            if header.startswith('name_'):
                runner_num = header.split('_')[1]
                if runner_num in winner_positions:
                    position = winner_positions[runner_num]
                    if position in POSITION_COLORS:
                        cell.fill = POSITION_COLORS[position]
                        cell.font = Font(bold=True)
                        cell.value = f"{value} (#{position})"
            col_idx += 1
        
        # Write duplicate detection columns
        ws.cell(row=row_idx, column=col_idx, value=duplicate_count)
        col_idx += 1
        ws.cell(row=row_idx, column=col_idx, value=last_seen)
    
    # Auto-fit columns
    for col in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for row in range(1, len(races) + 2):
            cell = ws.cell(row=row, column=col)
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add legend
    legend_row = len(races) + 4
    ws.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True)
    ws.cell(row=legend_row + 1, column=1, value="1st Place").fill = POSITION_COLORS[1]
    ws.cell(row=legend_row + 2, column=1, value="2nd Place").fill = POSITION_COLORS[2]
    ws.cell(row=legend_row + 3, column=1, value="3rd Place").fill = POSITION_COLORS[3]
    ws.cell(row=legend_row + 4, column=1, value="4th Place").fill = POSITION_COLORS[4]
    
    # Save workbook
    wb.save(OUTPUT_EXCEL)
    print(f"✅ Excel file updated: {OUTPUT_EXCEL}")
    print(f"   Races: {len(races)}, Results matched: {len(results)}")


class ResultsFileHandler(FileSystemEventHandler):
    """Handler for file system events on race_results.csv and races.csv."""
    
    def __init__(self):
        self.last_update = 0
        self.debounce_seconds = 2
    
    def on_modified(self, event):
        # Watch both results and races files
        if event.src_path.endswith('race_results.csv'):
            current_time = time.time()
            if current_time - self.last_update > self.debounce_seconds:
                self.last_update = current_time
                print(f"\n🏇 New results detected at {datetime.now().strftime('%H:%M:%S')}")
                create_excel_with_highlights()
        elif event.src_path.endswith('races.csv'):
            current_time = time.time()
            if current_time - self.last_update > self.debounce_seconds:
                self.last_update = current_time
                print(f"\n🏁 New race detected at {datetime.now().strftime('%H:%M:%S')}")
                create_excel_with_highlights()


def watch_for_results():
    """Watch for changes to race_results.csv and races.csv, update Excel."""
    print("=" * 60)
    print("Pattern Matcher - Watching for race data...")
    print("=" * 60)
    print(f"Watching: {RACES_CSV}")
    print(f"Watching: {RESULTS_CSV}")
    print(f"Output:   {OUTPUT_EXCEL}")
    print("\nPress Ctrl+C to stop\n")
    
    # Create initial Excel if races exist
    if RACES_CSV.exists():
        print("Creating initial Excel file...")
        create_excel_with_highlights()
    
    # Setup file watcher
    event_handler = ResultsFileHandler()
    observer = Observer()
    observer.schedule(event_handler, str(Path(__file__).parent), recursive=False)
    observer.start()
    
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("\n\nStopping pattern matcher...")
        observer.stop()
    
    observer.join()
    print("Pattern matcher stopped.")


if __name__ == "__main__":
    watch_for_results()
